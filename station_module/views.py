import datetime
import json
import os
import folium
from jalali_date import date2jalali
from openpyxl import Workbook
from django.db.models import Q
from folium.plugins import HeatMap, MarkerCluster
from openpyxl.utils import get_column_letter
from django.contrib.auth.decorators import login_required
from django.db.models import Count
from django.http import HttpRequest
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import ListView, DetailView
from openpyxl.styles import Font, Alignment, Border, PatternFill, Side
from site_module.models import SiteBanner
from utils.http_service import get_client_ip
from utils.convertors import group_list
from .models import StationCategory, StationType, StationVisit, StationGallery
from django.http import HttpResponse
from django.views import View
from django_tables2 import RequestConfig
from .tables import RecentRainGaugeTable, RecentRainGaugeLatLongTable
from django.shortcuts import render, redirect
from .models import Station
from django.contrib import messages
from .tables import Rainfall24hTable
from django.core.paginator import Paginator



class StationListView(ListView):
    template_name = 'station_module/station_list.html'
    model = Station
    context_object_name = 'stations'
    ordering = ['-code']
    paginate_by = 3

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super(StationListView, self).get_context_data()
        context['banners'] = SiteBanner.objects.filter(is_active=True, position__iexact='station_list')
        return context

    def get_queryset(self):
        query = super(StationListView, self).get_queryset()
        query = query.exclude(category__url_title__iexact='rain-gauge')
        category_name = self.kwargs.get('cat')
        type_name = self.kwargs.get('type')

        if type_name is not None:
            query = query.filter(type__url_title__iexact=type_name)
            print(query.query)

        if category_name is not None:
            query = query.filter(category__url_title__iexact=category_name)
            print(query.query)
        return query


class StationDetailView(DetailView):
    template_name = 'station_module/station_detail.html'
    model = Station

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        loaded_station = self.object
        request = self.request
        favorite_station_id = request.session.get('station_favorites')
        context['is_favorite'] = favorite_station_id == str(loaded_station.id)
        context['banners'] = SiteBanner.objects.filter(is_active=True, position__iexact='station_detail')
        galleries = list(StationGallery.objects.filter(station_id=loaded_station.id).all())
        galleries.insert(0, loaded_station)
        context['station_galleries_group'] = group_list(galleries, 3)
        context['related_stations'] = group_list(
            list(Station.objects.filter(type_id=loaded_station.type_id).exclude(pk=loaded_station.id).all()[:12]), 3)
        user_ip = get_client_ip(self.request)
        user_id = None
        if self.request.user.is_authenticated:
            user_id = self.request.user.id

        has_been_visited = StationVisit.objects.filter(ip__iexact=user_ip, station_id=loaded_station.id).exists()
        if not has_been_visited:
            new_visit = StationVisit(station=loaded_station, ip=user_ip, user_id=user_id)
            new_visit.save()
        return context


class AddStationFavorite(View):
    def post(self, request):
        station_id = request.POST['station_id']
        station = Station.objects.get(pk=station_id)
        request.session['station_favorites'] = station_id
        return redirect(station.get_absolute_url())


def station_categories_component(request: HttpRequest):
    station_categories = StationCategory.objects.annotate(stations_count=Count('station_categories')).filter(
        is_active=True, is_delete=False)
    context = {
        'categories': station_categories
    }
    # station_categories = StationCategory.objects.filter(is_active=True, is_delete=False)
    # context = {
    #     'categories': station_categories
    # }
    return render(request, 'station_module/components/station_categories_component.html', context)


def station_types_component(request: HttpRequest):
    station_types = StationType.objects.annotate(stations_count=Count('station')).filter(is_active=True)
    context = {
        'types': station_types
    }
    return render(request, 'station_module/components/station_type_component.html', context)


class RainGaugeTableView(View):
    template_name = 'station_module/rain_gauge_table.html'

    def get(self, request, *args, **kwargs):
        current_user = request.user
        if hasattr(current_user, 'employee'):
            workplace_code = ''
            workplace_code = current_user.employee.workplace_code
        else:
            print('no employee')
        rain_gauges = Station.objects.filter(parent_station__code__exact=workplace_code).order_by('-recent_rainfall')
        return render(request, self.template_name, {'rain_gauges': rain_gauges})


@csrf_exempt
def save_rain_gauge_value(request):
    id = request.POST.get('id')
    value = request.POST.get('value')

    rain_gauge = Station.objects.get(pk=id)
    rain_gauge.recent_rainfall = float(value)
    rain_gauge.last_rainfall_date_time = timezone.now()
    rain_gauge.save()

    return HttpResponse('Data successfully saved.')


def get_color(recent_rainfall):
    if recent_rainfall < 5:
        return 'yellow'
    elif 5 <= recent_rainfall < 10:
        return 'orange'
    elif 10 <= recent_rainfall < 30:
        return 'lightgreen'
    elif 30 <= recent_rainfall < 50:
        return 'green'
    elif 50 <= recent_rainfall < 80:
        return 'blue'
    else:
        return 'red'


@login_required
def rain_gauges_export_xls(request):
    today = timezone.now()
    period = int(request.GET.get('period', 24))  # مدت زمان انتخاب شده توسط کاربر (24 ساعت پیش فرض)
    start_date = today - datetime.timedelta(hours=period)  # محاسبه تاریخ شروع براساس مدت زمان انتخاب شده

    yesterday = today - datetime.timedelta(days=1)
    recent_rain = Station.objects.filter(last_rainfall_date_time__date__gte=start_date, recent_rainfall__gt=0).order_by(
        '-recent_rainfall')
    table = RecentRainGaugeTable(recent_rain)
    RequestConfig(request).configure(table)

    wb = Workbook()
    ws = wb.active

    # Set right-to-left direction
    ws.sheet_view.rightToLeft = True

    # Write table headers
    headers = ['ردیف', 'شهرستان', 'ایستگاه', 'مقدار بارندگی (میلیمتر)', 'زمان ثبت بارش']
    header_font = Font(bold=True, size=12, color='000000')
    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                    bottom=Side(style='thin'))

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Write table rows
    row_color = 'FFFFFF'  # Start with a white color for the first row
    for row_num, row in enumerate(table.rows, 2):
        ws.cell(row=row_num, column=1, value=row_num - 1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row_num, column=1).border = border  # Add border to the row number column
        for col_num, cell in enumerate(row, 2):
            excel_cell = ws.cell(row=row_num, column=col_num, value=str(cell))
            excel_cell.alignment = Alignment(horizontal='center', vertical='center')
            excel_cell.border = border

        # Apply row color
        if row_num % 2 == 0:
            for col_num in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col_num).fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6',
                                                                        fill_type='solid')  # Light gray
        else:
            for col_num in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col_num).fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF',
                                                                        fill_type='solid')  # White

    # Auto-size columns
    for col_num, column in enumerate(ws.columns, 1):
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[get_column_letter(col_num)].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=recent_rain.xlsx'
    wb.save(response)

    return response


def recent_rain_gauge(request):
    today = timezone.now()
    period = int(request.GET.get('period', 24))  # مدت زمان انتخاب شده توسط کاربر (24 ساعت پیش فرض)
    start_date = today - datetime.timedelta(hours=period)  # محاسبه تاریخ شروع براساس مدت زمان انتخاب شده

    # Filter stations with recent rainfall and exclude those without coordinates
    recent_rain_stations = Station.objects.filter(
        last_rainfall_date_time__gte=start_date
    ).exclude(longitude=None, latitude=None)

    # Create a map centered on Fars province
    m = folium.Map(location=[29.1044, 53.0454], zoom_start=7)  # Fars province coordinates

    # Load GeoJSON data for cities of Fars province borders
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    with open(os.path.join(BASE_DIR, 'geojson_files', 'FarsCity.json'), 'r') as f:
        geojson_data_cities = json.load(f)

    # Add GeoJSON layer for cities
    folium.GeoJson(
        geojson_data_cities,
        style_function=lambda feature: {
            'fillColor': '#ffffff00',  # Keep fill color transparent
            'color': 'gray',  # Use a different color for city borders, e.g., a shade of blue
            'weight': 2,  # Adjust border width for cities to be less than province borders
            'fillOpacity': 0.7,  # Adjust fill opacity as needed
        }
    ).add_to(m)

    # Load GeoJSON data for Fars province borders
    with open(os.path.join(BASE_DIR, 'geojson_files', 'Fars.json'), 'r') as f:
        geojson_data_province = json.load(f)

    # Add GeoJSON layer for province
    folium.GeoJson(
        geojson_data_province,
        style_function=lambda feature: {
            'fillColor': '#ffffff00',  # Transparent fill color
            'color': 'black',  # Border color
            'weight': 3,  # Border width
            'fillOpacity': 0.5,
        }
    ).add_to(m)

    # Create a list of [latitude, longitude, weight] for the heatmap
    heat_data = [[station.latitude, station.longitude, station.recent_rainfall] for station in recent_rain_stations if
                 station.recent_rainfall > 0]

    # Add heatmap layer
    HeatMap(heat_data).add_to(m)

    # Add marker clustering
    marker_cluster = MarkerCluster().add_to(m)

    # Add markers to the cluster
    for station in recent_rain_stations:
        if station.recent_rainfall > 0:
            folium.Marker(
                location=[station.latitude, station.longitude],
                popup=f"{station.title} - بارندگی اخیر: {station.recent_rainfall} میلیمتر",
                icon=folium.Icon(color=get_color(station.recent_rainfall), icon='cloud')
            ).add_to(marker_cluster)

    # Convert folium map to HTML string
    map_html = m._repr_html_()

    # Fetch recent rainfall data for table
    recent_rain = Station.objects.filter(
        last_rainfall_date_time__gte=start_date
    ).order_by('-recent_rainfall')

    # Create and configure table
    table = RecentRainGaugeTable(recent_rain)
    RequestConfig(request).configure(table)

    allowed_users = ['2559110393', '2450451331']

    return render(request, 'station_module/recent_rain_gauge.html',
                  {'table': table, 'today': today, 'map_html': map_html, 'allowed_users': allowed_users,
                   'period': period})


@login_required
def rain_gauges_export_lat_long_xls(request):
    period = int(request.GET.get('period', 24))  # مدت زمان انتخاب شده توسط کاربر (24 ساعت پیش فرض)
    today = timezone.now()
    start_date = today - datetime.timedelta(hours=period)

    recent_rain = Station.objects.filter(last_rainfall_date_time__gte=start_date).order_by('-recent_rainfall')
    table = RecentRainGaugeLatLongTable(recent_rain)
    RequestConfig(request).configure(table)

    wb = Workbook()
    ws = wb.active

    # Set right-to-left direction
    ws.sheet_view.rightToLeft = True

    # Write table headers
    headers = ['ردیف', 'شهرستان', 'ایستگاه', 'عرض ج', 'طول ج', 'مقدار بارندگی (میلیمتر)', 'زمان ثبت بارش']
    header_font = Font(bold=True, size=12, color='000000')
    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                    bottom=Side(style='thin'))

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Write table rows
    row_color = 'FFFFFF'  # Start with a white color for the first row
    for row_num, row in enumerate(table.rows, 2):
        ws.cell(row=row_num, column=1, value=row_num - 1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row_num, column=1).border = border  # Add border to the row number column
        for col_num, cell in enumerate(row, 2):
            excel_cell = ws.cell(row=row_num, column=col_num, value=str(cell))
            excel_cell.alignment = Alignment(horizontal='center', vertical='center')
            excel_cell.border = border  # Add border to the remaining columns

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Create an HTTP response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=rain_gauges_lat_long.xlsx'
    wb.save(response)
    return response


@login_required
def table_update_rainfall(request):
    current_user = request.user
    workplace_code = None
    if hasattr(current_user, 'employee'):
        workplace_code = current_user.employee.workplace_code

    stations = Station.objects.filter(parent_station__code__exact=workplace_code, code__gte=19000, code__lte=99999)

    if request.method == 'POST':
        for station in stations:
            # دریافت مقادیر از فرم با استفاده از شناسه ایستگاه
            rainfall_3h = request.POST.get(f'rainfall_3h_{station.id}', '').strip()
            recent_rainfall = request.POST.get(f'recent_rainfall_{station.id}', '').strip()
            year_rainfall = request.POST.get(f'year_rainfall_{station.id}', '').strip()

            # تبدیل مقادیر دریافتی به float و بررسی اینکه آیا مقدار خالی یا منفی نباشد
            try:
                if rainfall_3h and float(rainfall_3h) >= 0:
                    station.rainfall_3h = float(rainfall_3h)
                if recent_rainfall and float(recent_rainfall) >= 0:
                    station.recent_rainfall = float(recent_rainfall)
                if year_rainfall and float(year_rainfall) >= 0:
                    station.year_rainfall = float(year_rainfall)

                # اگر حداقل یکی از فیلدها مقدار دارد، مدل را ذخیره کنید
                if rainfall_3h or recent_rainfall or year_rainfall:
                    station.save()
            except ValueError:
                # اگر مقدار نمی‌تواند به float تبدیل شود یا منفی است، خطا را مدیریت کنید
                pass

    return render(request, 'station_module/table_update_rainfall.html', {'stations': stations})


def stations_table_3h_recent_rainfall(request):
    # تعریف زمان‌های امروز و دیروز
    today = timezone.now()
    yesterday = today - datetime.timedelta(days=1)

    # ایجاد کوئری با استفاده از Q objects برای ترکیب شرایط
    stations = Station.objects.filter(
        Q(last_rainfall_date_time__date__gte=yesterday) &
        Q(code__gte=19000, code__lte=99999)
    ).order_by('-recent_rainfall')

    # اگر درخواست برای دانلود فایل اکسل باشد
    if 'download' in request.GET:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="stations.xlsx"'

        wb = Workbook()
        ws = wb.active
        ws.title = "Stations"

        # تنظیم جهت صفحه از راست به چپ
        ws.sheet_view.rightToLeft = True

        # نوشتن سربرگ‌های جدول
        columns = ['ردیف', 'نام ایستگاه', 'بارندگی 3 ساعته', 'جمع بارندگی سامانه اخیر (میلیمتر)',
                   'جمع بارندگی سال زرعی تا کنون (میلیمتر)']
        header_font = Font(bold=True, size=12, color='000000')
        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                        bottom=Side(style='thin'))

        for col_num, column_title in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num, value=column_title)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # نوشتن سطرهای جدول
        for row_num, station in enumerate(stations, 2):
            ws.cell(row=row_num, column=1, value=row_num - 1).alignment = Alignment(horizontal='center',
                                                                                    vertical='center')
            ws.cell(row=row_num, column=1).border = border
            ws.cell(row=row_num, column=2, value=station.title).alignment = Alignment(horizontal='center',
                                                                                      vertical='center')
            ws.cell(row=row_num, column=2).border = border
            ws.cell(row=row_num, column=3, value=station.rainfall_3h).alignment = Alignment(horizontal='center',
                                                                                            vertical='center')
            ws.cell(row=row_num, column=3).border = border
            ws.cell(row=row_num, column=4, value=station.recent_rainfall).alignment = Alignment(horizontal='center',
                                                                                                vertical='center')
            ws.cell(row=row_num, column=4).border = border
            ws.cell(row=row_num, column=5, value=station.year_rainfall).alignment = Alignment(horizontal='center',
                                                                                              vertical='center')
            ws.cell(row=row_num, column=5).border = border
            # ...

            # اعمال رنگ سطر
            row_fill = PatternFill(start_color='FFFFFF' if row_num % 2 == 1 else 'E6E6E6',
                                   end_color='FFFFFF' if row_num % 2 == 1 else 'E6E6E6', fill_type='solid')
            for col_num in range(1, len(columns) + 1):
                ws.cell(row=row_num, column=col_num).fill = row_fill

        # تنظیم اندازه خودکار ستون‌ها
        for col_num, column in enumerate(ws.columns, 1):
            max_length = max(len(str(cell.value)) for cell in column)
            adjusted_width = (max_length + 0) * 1.0
            ws.column_dimensions[get_column_letter(col_num)].width = adjusted_width

        wb.save(response)
        return response

    # اگر درخواست برای نمایش صفحه باشد
    return render(request, 'station_module/stations_table_3h_recent_rainfall.html', {'stations': stations})


@login_required
def table_update_24h_rainfall(request):
    current_user = request.user
    workplace_code = None
    if hasattr(current_user, 'employee'):
        workplace_code = current_user.employee.workplace_code

    stations = Station.objects.filter(
        parent_station__code__exact=workplace_code)

    if request.method == 'POST':
        has_errors = False
        for station in stations:
            # دریافت مقادیر از فرم
            rainfall_24h = request.POST.get(f'rainfall_24h_{station.id}', '').strip()
            recent_rainfall = request.POST.get(f'recent_rainfall_{station.id}', '').strip()
            year_rainfall = request.POST.get(f'year_rainfall_{station.id}', '').strip()

            # اگر همه فیلدها خالی هستند، رد شود
            if not (rainfall_24h or recent_rainfall or year_rainfall):
                continue

            # بررسی اینکه آیا همه فیلدهای مربوط به یک ایستگاه پر شده‌اند
            if not (rainfall_24h and recent_rainfall and year_rainfall):
                messages.error(request, f'برای ایستگاه {station.title} باید همه مقادیر بارش وارد شوند')
                has_errors = True
                continue

            try:
                # تبدیل و اعتبارسنجی مقادیر
                rainfall_24h_float = float(rainfall_24h)
                recent_rainfall_float = float(recent_rainfall)
                year_rainfall_float = float(year_rainfall)

                # بررسی مقادیر منفی
                if rainfall_24h_float < 0 or recent_rainfall_float < 0 or year_rainfall_float < 0:
                    messages.error(request, f'مقادیر بارش برای ایستگاه {station.title} نباید منفی باشند')
                    has_errors = True
                    continue

                # ذخیره اطلاعات
                station.rainfall_24h = rainfall_24h_float
                station.recent_rainfall = recent_rainfall_float
                station.year_rainfall = year_rainfall_float
                station.save()

            except ValueError:
                messages.error(request, f'خطا در ورودی برای ایستگاه {station.title}: مقادیر باید عددی باشند')
                has_errors = True

        # اگر خطایی وجود داشته باشد
        if has_errors:
            return render(request, 'station_module/table_update_24h_rainfall.html', {'stations': stations})
        else:
            messages.success(request, 'اطلاعات با موفقیت ذخیره شد')
            return redirect('table-update-24h-rainfall')

    return render(request, 'station_module/table_update_24h_rainfall.html', {'stations': stations})





def rainfall_24h(request):
    today = timezone.now()

    # دریافت بازه زمانی مورد نظر (24، 48، 72، یا 96 ساعت)
    period = int(request.GET.get('period', 12))
    rainfall_type = request.GET.get('rainfall_type', 'recent')  # نوع بارش (روزانه، اخیر، زراعی)

    # تبدیل تاریخ به جلالی (فرض می‌شود تابع date2jalali تعریف شده است)
    date_jalali = date2jalali(today - datetime.timedelta(days=365)).strftime('%d-%m-%Y')

    # محاسبه بازه‌های زمانی
    start_date_24h = today - datetime.timedelta(hours=12)
    start_date_recent = today - datetime.timedelta(days=4)

    # دریافت داده‌ها و مقداردهی صفر به ایستگاه‌های بدون دیتا یا با داده‌های قدیمی
    stations = Station.objects.all()
    for station in stations:
        # صفر کردن بارش 24 ساعته اگر از 12 ساعت گذشته باشد
        if not station.last_rainfall_date_time or station.last_rainfall_date_time < start_date_24h:
            station.rainfall_24h = 0
            station.save()

        # صفر کردن بارش اخیر اگر از 4 روز گذشته باشد
        if not station.last_rainfall_date_time or station.last_rainfall_date_time < start_date_recent:
            station.recent_rainfall = 0
            station.save()

    # فیلتر براساس نوع بارش
    if rainfall_type == 'daily':
        stations = stations.order_by('-rainfall_24h')
    elif rainfall_type == 'recent':
        stations = stations.order_by('-recent_rainfall')
    elif rainfall_type == 'crop':
        stations = stations.order_by('-year_rainfall')

    # تنظیم صفحه‌بندی
    paginator = Paginator(stations, 20)
    page_number = request.GET.get('page', 1)
    page = paginator.get_page(page_number)

    # محاسبه شماره ردیف برای صفحه جاری
    start_row_number = (page.number - 1) * paginator.per_page

    # ایجاد جدول
    table = Rainfall24hTable(stations, start_row_number=start_row_number)
    RequestConfig(request, paginate={"per_page": 20}).configure(table)

    # ساخت نقشه
    m = folium.Map(location=[29.1044, 53.0454], zoom_start=7)

    # بارگذاری فایل GeoJSON برای مرز شهرها
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    with open(os.path.join(BASE_DIR, 'geojson_files', 'FarsCity.json'), 'r') as f:
        geojson_data_cities = json.load(f)

    folium.GeoJson(
        geojson_data_cities,
        style_function=lambda feature: {
            'fillColor': '#ffffff00',
            'color': 'gray',
            'weight': 2,
            'fillOpacity': 0.7,
        }
    ).add_to(m)

    # بارگذاری فایل GeoJSON برای مرز استان
    with open(os.path.join(BASE_DIR, 'geojson_files', 'Fars.json'), 'r') as f:
        geojson_data_province = json.load(f)

    folium.GeoJson(
        geojson_data_province,
        style_function=lambda feature: {
            'fillColor': '#ffffff00',
            'color': 'black',
            'weight': 3,
            'fillOpacity': 0.5,
        }
    ).add_to(m)

    # ایجاد داده‌های HeatMap
    if rainfall_type == 'daily':
        heat_data = [
            [station.latitude, station.longitude, station.rainfall_24h]
            for station in stations if station.rainfall_24h is not None and station.rainfall_24h > 0
        ]
    elif rainfall_type == 'recent':
        heat_data = [
            [station.latitude, station.longitude, station.recent_rainfall]
            for station in stations if station.recent_rainfall is not None and station.recent_rainfall > 0
        ]
    elif rainfall_type == 'crop':
        heat_data = [
            [station.latitude, station.longitude, station.year_rainfall]
            for station in stations if station.year_rainfall is not None and station.year_rainfall > 0
        ]

    HeatMap(heat_data).add_to(m)

    # اضافه کردن MarkerCluster
    marker_cluster = MarkerCluster().add_to(m)

    for station in stations:
        rainfall_value = 0
        if rainfall_type == 'daily':
            rainfall_value = station.rainfall_24h
        elif rainfall_type == 'recent':
            rainfall_value = station.recent_rainfall
        elif rainfall_type == 'crop':
            rainfall_value = station.year_rainfall

        # فقط اگر مقدار بارش معتبر باشد، مارکر اضافه می‌شود
        if rainfall_value is not None and rainfall_value > 0:
            folium.Marker(
                location=[station.latitude, station.longitude],
                popup=f"{station.title} - بارندگی: {rainfall_value} میلیمتر",
                icon=folium.Icon(color="blue", icon='cloud')
            ).add_to(marker_cluster)

    # تبدیل نقشه به HTML
    map_html = m._repr_html_()

    # کاربران مجاز
    allowed_users = ['2559110393', '2450451331']

    return render(request, 'station_module/rainfall_24h.html', {
        'table': table,
        'rainfall_type': rainfall_type,
        'period': period,
        'map_html': map_html,
        'today': today,
        'date_jalali': date_jalali,
        'allowed_users': allowed_users
    })

# def rainfall_24h(request):
#     today = timezone.now()
#
#     # دریافت بازه زمانی مورد نظر (24، 48، 72، یا 96 ساعت)
#     period = int(request.GET.get('period', 12))
#     rainfall_type = request.GET.get('rainfall_type', 'recent')  # نوع بارش (روزانه، اخیر، زراعی)
#
#     date_jalali = date2jalali(today - datetime.timedelta(days=365)).strftime('%d-%m-%Y')
#
#     # محاسبه بازه‌های زمانی
#     start_date_24h = today - datetime.timedelta(hours=12)  # تغییر به 12 ساعت
#     start_date_recent = today - datetime.timedelta(days=4)
#
#     # دریافت داده‌ها و مقداردهی صفر به ایستگاه‌های بدون دیتا یا با داده‌های قدیمی
#     stations = Station.objects.all()
#     for station in stations:
#         # صفر کردن بارش 24 ساعته اگر از 12 ساعت گذشته باشد
#         if not station.last_rainfall_date_time or station.last_rainfall_date_time < start_date_24h:
#             station.rainfall_24h = 0
#             station.save()  # ذخیره تغییرات در مدل
#
#         # صفر کردن بارش اخیر اگر از 4 روز گذشته باشد
#         if not station.last_rainfall_date_time or station.last_rainfall_date_time < start_date_recent:
#             station.recent_rainfall = 0
#             station.save()  # ذخیره تغییرات در مدل
#
#     # فیلتر براساس نوع بارش
#     if rainfall_type == 'daily':
#         stations = stations.order_by('-rainfall_24h')  # مرتب‌سازی بر اساس بارش 24 ساعته از بیشترین به کمترین
#     elif rainfall_type == 'recent':
#         stations = stations.order_by('-recent_rainfall')  # مرتب‌سازی بر اساس بارش اخیر از بیشترین به کمترین
#     elif rainfall_type == 'crop':
#         stations = stations.order_by('-year_rainfall')  # مرتب‌سازی بر اساس بارش سال زراعی از بیشترین به کمترین
#
#     # ایجاد جدول
#     table = Rainfall24hTable(stations)
#     RequestConfig(request, paginate={"per_page": 20}).configure(table)
#
#     # ساخت نقشه
#     m = folium.Map(location=[29.1044, 53.0454], zoom_start=7)  # مختصات استان فارس
#
#     # بارگذاری فایل GeoJSON برای مرز شهرها
#     BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
#     with open(os.path.join(BASE_DIR, 'geojson_files', 'FarsCity.json'), 'r') as f:
#         geojson_data_cities = json.load(f)
#
#     folium.GeoJson(
#         geojson_data_cities,
#         style_function=lambda feature: {
#             'fillColor': '#ffffff00',
#             'color': 'gray',
#             'weight': 2,
#             'fillOpacity': 0.7,
#         }
#     ).add_to(m)
#
#     # بارگذاری فایل GeoJSON برای مرز استان
#     with open(os.path.join(BASE_DIR, 'geojson_files', 'Fars.json'), 'r') as f:
#         geojson_data_province = json.load(f)
#
#     folium.GeoJson(
#         geojson_data_province,
#         style_function=lambda feature: {
#             'fillColor': '#ffffff00',
#             'color': 'black',
#             'weight': 3,
#             'fillOpacity': 0.5,
#         }
#     ).add_to(m)
#
#     # ایجاد داده‌های HeatMap
#     if rainfall_type == 'daily':
#         heat_data = [[station.latitude, station.longitude, station.rainfall_24h] for station in stations if
#                      station.rainfall_24h > 0]
#     elif rainfall_type == 'recent':
#         heat_data = [[station.latitude, station.longitude, station.recent_rainfall] for station in stations if
#                      station.recent_rainfall > 0]
#     elif rainfall_type == 'crop':
#         heat_data = [[station.latitude, station.longitude, station.year_rainfall] for station in stations if
#                      station.year_rainfall > 0]
#
#     HeatMap(heat_data).add_to(m)
#
#     # اضافه کردن MarkerCluster
#     marker_cluster = MarkerCluster().add_to(m)
#
#     for station in stations:
#         rainfall_value = 0
#         if rainfall_type == 'daily':
#             rainfall_value = station.rainfall_24h
#         elif rainfall_type == 'recent':
#             rainfall_value = station.recent_rainfall
#         elif rainfall_type == 'crop':
#             rainfall_value = station.year_rainfall
#
#         if rainfall_value > 0:
#             folium.Marker(
#                 location=[station.latitude, station.longitude],
#                 popup=f"{station.title} - بارندگی: {rainfall_value} میلیمتر",
#                 icon=folium.Icon(color="blue", icon='cloud')
#             ).add_to(marker_cluster)
#
#     # تبدیل نقشه به HTML
#     map_html = m._repr_html_()
#     allowed_users = ['2559110393', '2450451331']
#
#
#     return render(request, 'station_module/rainfall_24h.html', {
#         'table': table,
#         'rainfall_type': rainfall_type,
#         'period': period,
#         'map_html': map_html,
#         'today': today,
#         'date_jalali': date_jalali,
#         'allowed_users': allowed_users
#     })


@login_required
def rainfall_24h_export_xls(request):
    today = timezone.now()

    # دریافت بازه زمانی مورد نظر (24، 48، 72، یا 96 ساعت)
    period = int(request.GET.get('period', 12))
    rainfall_type = request.GET.get('rainfall_type', 'daily')  # نوع بارش (روزانه، اخیر، زراعی)

    date_jalali = date2jalali(today - datetime.timedelta(days=365)).strftime('%d-%m-%Y')

    # محاسبه بازه‌های زمانی
    start_date_24h = today - datetime.timedelta(hours=12)  # تغییر به 12 ساعت
    start_date_recent = today - datetime.timedelta(days=4)

    # دریافت داده‌ها و مقداردهی صفر به ایستگاه‌های بدون دیتا یا با داده‌های قدیمی
    stations = Station.objects.all().order_by('-recent_rainfall')
    for station in stations:
        # صفر کردن بارش 24 ساعته اگر از 12 ساعت گذشته باشد
        if not station.last_rainfall_date_time or station.last_rainfall_date_time < start_date_24h:
            station.rainfall_24h = 0
            station.save()  # ذخیره تغییرات در مدل

        # صفر کردن بارش اخیر اگر از 4 روز گذشته باشد
        if not station.last_rainfall_date_time or station.last_rainfall_date_time < start_date_recent:
            station.recent_rainfall = 0
            station.save()  # ذخیره تغییرات در مدل

    # فیلتر براساس نوع بارش
    if rainfall_type == 'daily':
        stations = stations  # مرتب‌سازی بر اساس بارش 24 ساعته از بیشترین به کمترین
    elif rainfall_type == 'recent':
        stations = stations.order_by('-recent_rainfall')  # مرتب‌سازی بر اساس بارش اخیر از بیشترین به کمترین
    elif rainfall_type == 'crop':
        stations = stations  # مرتب‌سازی بر اساس بارش سال زراعی از بیشترین به کمترین

    # ایجاد فایل اکسل
    wb = Workbook()
    ws = wb.active
    ws.sheet_view.rightToLeft = True

    ws.title = "بارش‌ها"

    # تنظیمات استایل
    header_font = Font(bold=True, size=12, color='000000')
    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                    bottom=Side(style='thin'))

    # نوشتن هدرها
    headers = ['ردیف', 'شهرستان', 'ایستگاه', 'بارش 24 ساعته (میلیمتر)', 'بارش اخیر (میلیمتر)', 'بارش سال آبی (میلیمتر)',
               'تاریخ و ساعت ثبت']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        cell.fill = header_fill

    # نوشتن داده‌ها
    for row_num, station in enumerate(stations, 2):
        ws.cell(row=row_num, column=1, value=row_num - 1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row_num, column=1).border = border  # افزودن مرز به ستون ردیف
        ws.cell(row=row_num, column=2, value=station.city).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row_num, column=3, value=station.title).alignment = Alignment(horizontal='center',
                                                                                  vertical='center')
        ws.cell(row=row_num, column=4, value=station.rainfall_24h).alignment = Alignment(horizontal='center',
                                                                                         vertical='center')
        ws.cell(row=row_num, column=5, value=station.recent_rainfall).alignment = Alignment(horizontal='center',
                                                                                            vertical='center')
        ws.cell(row=row_num, column=6, value=station.year_rainfall).alignment = Alignment(horizontal='center',
                                                                                          vertical='center')
        ws.cell(row=row_num, column=7, value=station.last_rainfall_date_time.strftime(
            '%H:%M:%S - %Y/%m/%d') if station.last_rainfall_date_time else '').alignment = Alignment(
            horizontal='center', vertical='center')

        # افزودن مرز به سلول‌ها
        for col_num in range(1, 8):
            ws.cell(row=row_num, column=col_num).border = border

        # رنگ‌بندی سطرها
        if row_num % 2 == 0:
            for col_num in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col_num).fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6',
                                                                        fill_type='solid')
        else:
            for col_num in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col_num).fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF',
                                                                        fill_type='solid')

    # تنظیم عرض ستون‌ها
    for col_num, column in enumerate(ws.columns, 1):
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[get_column_letter(col_num)].width = adjusted_width

    # ایجاد پاسخ دانلود اکسل
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=rainfall_24h_data_{today.strftime("%Y-%m-%d")}.xlsx'
    wb.save(response)

    return response


@login_required
def rainfall_24h_export_lat_long_xls(request):
    today = timezone.now()

    # دریافت بازه زمانی مورد نظر (24، 48، 72، یا 96 ساعت)
    period = int(request.GET.get('period', 12))
    rainfall_type = request.GET.get('rainfall_type', 'daily')  # نوع بارش (روزانه، اخیر، زراعی)

    date_jalali = date2jalali(today - datetime.timedelta(days=365)).strftime('%d-%m-%Y')

    # محاسبه بازه‌های زمانی
    start_date_24h = today - datetime.timedelta(hours=12)  # تغییر به 12 ساعت
    start_date_recent = today - datetime.timedelta(days=4)

    # دریافت داده‌ها و مقداردهی صفر به ایستگاه‌های بدون دیتا یا با داده‌های قدیمی
    stations = Station.objects.all().order_by('-recent_rainfall')
    for station in stations:
        # صفر کردن بارش 24 ساعته اگر از 12 ساعت گذشته باشد
        if not station.last_rainfall_date_time or station.last_rainfall_date_time < start_date_24h:
            station.rainfall_24h = 0
            station.save()  # ذخیره تغییرات در مدل

        # صفر کردن بارش اخیر اگر از 4 روز گذشته باشد
        if not station.last_rainfall_date_time or station.last_rainfall_date_time < start_date_recent:
            station.recent_rainfall = 0
            station.save()  # ذخیره تغییرات در مدل

    # فیلتر براساس نوع بارش
    if rainfall_type == 'daily':
        stations = stations  # مرتب‌سازی بر اساس بارش 24 ساعته از بیشترین به کمترین
    elif rainfall_type == 'recent':
        stations = stations.order_by('-recent_rainfall')  # مرتب‌سازی بر اساس بارش اخیر از بیشترین به کمترین
    elif rainfall_type == 'crop':
        stations = stations  # مرتب‌سازی بر اساس بارش سال زراعی از بیشترین به کمترین

    # ایجاد فایل اکسل
    wb = Workbook()
    ws = wb.active
    ws.sheet_view.rightToLeft = True

    ws.title = "بارش‌ها"

    # تنظیمات استایل
    header_font = Font(bold=True, size=12, color='000000')
    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                    bottom=Side(style='thin'))

    # نوشتن هدرها
    headers = ['ردیف', 'شهرستان', 'ایستگاه', 'عرض ج', 'طول ج', 'بارش 24 ساعته (میلیمتر)', 'بارش اخیر (میلیمتر)',
               'بارش سال آبی (میلیمتر)',
               'تاریخ و ساعت ثبت']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        cell.fill = header_fill

    # نوشتن داده‌ها
    for row_num, station in enumerate(stations, 2):
        ws.cell(row=row_num, column=1, value=row_num - 1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row_num, column=1).border = border  # افزودن مرز به ستون ردیف
        ws.cell(row=row_num, column=2, value=station.city).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row_num, column=3, value=station.title).alignment = Alignment(horizontal='center',
                                                                                  vertical='center')
        ws.cell(row=row_num, column=4, value=station.latitude).alignment = Alignment(horizontal='center',
                                                                                     vertical='center')
        ws.cell(row=row_num, column=5, value=station.longitude).alignment = Alignment(horizontal='center',
                                                                                      vertical='center')
        ws.cell(row=row_num, column=6, value=station.rainfall_24h).alignment = Alignment(horizontal='center',
                                                                                         vertical='center')
        ws.cell(row=row_num, column=7, value=station.recent_rainfall).alignment = Alignment(horizontal='center',
                                                                                            vertical='center')
        ws.cell(row=row_num, column=8, value=station.year_rainfall).alignment = Alignment(horizontal='center',
                                                                                          vertical='center')
        ws.cell(row=row_num, column=9, value=station.last_rainfall_date_time.strftime(
            '%H:%M:%S - %Y/%m/%d') if station.last_rainfall_date_time else '').alignment = Alignment(
            horizontal='center', vertical='center')

        # افزودن مرز به سلول‌ها
        for col_num in range(1, 10):
            ws.cell(row=row_num, column=col_num).border = border

        # رنگ‌بندی سطرها
        if row_num % 2 == 0:
            for col_num in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col_num).fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6',
                                                                        fill_type='solid')
        else:
            for col_num in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col_num).fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF',
                                                                        fill_type='solid')

    # تنظیم عرض ستون‌ها
    for col_num, column in enumerate(ws.columns, 1):
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[get_column_letter(col_num)].width = adjusted_width

    # ایجاد پاسخ دانلود اکسل
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=rainfall_24h_data_{today.strftime("%Y-%m-%d")}.xlsx'
    wb.save(response)

    return response
